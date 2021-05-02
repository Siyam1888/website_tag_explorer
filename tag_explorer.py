import subprocess
import json
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


class WebsiteTagExplorer:
    """Run the website-evidence-collector and processes the output"""

    def __init__(self):
        self.command = "website-evidence-collector --json --quiet --no-output"

    def get_output(self, url):
        """Run the website-evidence-collector and returns the output as a dictionary"""
        process = subprocess.run(
            f"{self.command} {url}", shell=True, capture_output=True
        )
        if process.returncode == 0:
            json_output = process.stdout.decode()
            output_dict = json.loads(json_output)

            return output_dict

    def get_hosts(self, output):
        """Read the hosts and return the unique ones"""
        hosts = output["hosts"]
        return_values = set()

        # iterating over the keys of hosts e.g requests, beacons, cookies, links
        for host_category in hosts:
            # iterting over the keys of each host e.g firstParty, thirdParty
            for link_type in hosts[host_category]:
                # iterating over links of each link type e.g www.gsk.com
                for link in hosts[host_category][link_type]:
                    # value = {'url': url, 'host_category': host_category, 'link_type': link_type, 'link': link}
                    # return_values.append(value)

                    # adding only third party links
                    if link_type == "thirdParty":
                        return_values.add(link)

        return return_values

    def get_beacons(self, output):
        """Read the beacons from the given website-evidence-collector output and return a set"""
        beacons = output["beacons"]
        return_values = set()

        for beacon in beacons:
            return_values.add(beacon["url"])

        return return_values


class Excel:
    """Work with all excel related tasks."""

    def __init__(self, file_name):
        self.file_name = file_name
        self.font = Font(color="000000", bold=True)
        self.bg_color = PatternFill(fgColor="E8E8E8", fill_type="solid")
        self.customize_excel()

    def create_sheets(self):
        """Create all the sheets required for the project."""
        self.tags = (
            self.wb.create_sheet("Tags")
            if "Tags" not in self.wb.sheetnames
            else self.wb["Tags"]
        )
        self.errors = (
            self.wb.create_sheet("Errors")
            if "Errors" not in self.wb.sheetnames
            else self.wb["Errors"]
        )

    def make_columns(self, cells_zip, sheet, general_width=20, url_width=50):
        """Takes zip values of rows and columns and puts values in place with some stylings"""
        # iterating through the column and its values to put them in place
        for col, value in cells_zip:
            cell = sheet[f"{col}1"]
            cell.value = value
            cell.font = self.font
            cell.fill = self.bg_color
            sheet.freeze_panes = cell

            # fixing the column width
            sheet.column_dimensions[col].width = general_width
        # fixing the URL column width
        sheet.column_dimensions["A"].width = url_width

    def customize_tags_column(self):
        """Customize the Tags column according to its values"""

        # combining columns with its values
        tags_column = zip(
            ("A", "B", "C", "D"),
            (
                "Site URL",
                "Tag Name",
                "Tag URL",
            ),
        )
        self.make_columns(tags_column, self.tags)

    def customize_excel(self):
        """Run all the functions related to excel customization"""
        self.wb = load_workbook(self.file_name)
        self.create_sheets()
        self.customize_tags_column()
        self.wb.save(self.file_name)

    def generate_inputs(self):
        """Read the first column of Input sheet and yield the values"""
        inputs = self.wb["Input"]
        for row in range(2, inputs.max_row + 1):
            # generates the links one by one
            if value := inputs[f"A{row}"].value:
                yield value

    def append_tags(self, tags):
        """Read the hosts, host categories, link types, links and append them to excel"""

        for host in tags:
            self.wb["Tags"].append(
                (
                    host["site"],
                    host["tag_name"],
                    host["tag_url"],
                )
            )

            self.wb.save(self.file_name)


def get_final_output(url):
    """Takes an URL and returns all the tags found on the url -> tag_name, tag_url"""
    tag_explorer = WebsiteTagExplorer()

    # Getting the output from the website evidence collector
    output = tag_explorer.get_output(url)
    if output:
        beacons = tag_explorer.get_beacons(output)
        hosts = tag_explorer.get_hosts(output)
        # Loading the list of tag names and corresponding urls from the json
        with open("tags_list.json") as f:
            tags_list = json.load(f)

        # list of dictionaries to return
        tags_found = []
        named_tags = set()
        found_beacons = set()

        # Naming the tags from url and adding them to the list
        for tag in tags_list:

            for tag_url in tag["tag_urls"]:
                # matching beacon urls
                for beacon_url in beacons:
                    if tag_url in beacon_url:
                        tags_found.append(
                            {
                                "site": url,
                                "tag_name": tag["tag_name"],
                                "tag_url": beacon_url,
                            }
                        )
                        found_beacons.add(tag["tag_name"])

                # matching hosts
                for host in hosts:
                    if (host in tag_url) or (tag_url in host):
                        # print(tag['tag_name'], url, host)

                        # if tag['tag_name'] not in found_beacons:
                        tags_found.append(
                            {"site": url, "tag_name": tag["tag_name"], "tag_url": host}
                        )
                        named_tags.add(host)

        # Adding the unnamed tags
        for host in hosts - named_tags:
            tags_found.append({"site": url, "tag_name": None, "tag_url": host})

        return tags_found

    print(f"{url} : The evidence collector did not return anything :(")


def main():
    filename = "Output.xlsx"
    excel = Excel(filename)
    for url in excel.generate_inputs():
        tags = get_final_output(url)
        # print(tags)
        if tags:
            excel.append_tags(tags)
        else:
            excel.errors.append((url,))

    excel.wb.save(filename)


main()
