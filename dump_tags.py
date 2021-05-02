from openpyxl import load_workbook
import json
import webbrowser


file_path = "Tags.xlsx"
wb = load_workbook(file_path)


def dump_tags():
    """Read the tag names and tag urls and save the list of their pairs as json"""
    inputs = wb["Input"]

    tags_list = []
    for row in range(2, inputs.max_row + 1):
        # generates the links one by one
        tag_name = inputs[f"A{row}"].value
        tag_url = inputs[f"C{row}"].value
        if tag_name and tag_url:
            tag = {
                "tag_name": tag_name,
                "tag_urls": list(
                    filter(
                        lambda x: x != "",
                        map(
                            lambda x: x.replace("https://", "").strip(),
                            tag_url.strip().split(","),
                        ),
                    )
                ),
            }
            tags_list.append(tag)

    with open("tags_list.json", "w") as f:
        json.dump(tags_list, f)


dump_tags()
